"""Phase EXCEL_EXPORT - Export Excel formaté professionnel.

Basé sur csv_to_excel_formatted.py
Génère un fichier Excel lisible avec mise en forme professionnelle.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass
from typing import Optional, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


@dataclass
class ExcelExportConfig:
    """Configuration pour l'export Excel."""
    enabled: bool = False
    output_excel: Optional[Path] = None


def infer_statut(row) -> str:
    """Déduit le statut du marché à partir de la date limite."""
    date_limite = row.get('date_limite_remise_offres')
    if pd.isna(date_limite) or date_limite == '':
        return 'INCONCLUSIVE'
    
    try:
        if isinstance(date_limite, str):
            date_limite = pd.to_datetime(date_limite)
        
        if date_limite < datetime.now():
            return 'CLOTUREE'
        else:
            return 'OUVERTE'
    except:
        return 'INCONCLUSIVE'


def get_column_widths(df: pd.DataFrame) -> Dict[str, int]:
    """Définit les largeurs de colonnes personnalisées."""
    custom_widths = {
        'reference': 18,
        'titre': 50,
        'acheteur': 35,
        'type_acheteur': 18,
        'fonction_publique': 18,
        'procedure_type': 15,
        'famille_procedure_deduite': 20,
        'typologie_marche_verifiee': 15,
        'type_marche': 12,
        'ccag_type': 22,
        'cpv_principal': 15,
        'cpv_secondaires': 35,
        'localisation': 25,
        'date_limite_remise_offres': 22,
        'duree': 10,
        'renouvellements': 30,
        'montant_estime': 18,
        'seuil_formalise_applicable': 18,
        'url_marche': 50,
        'url_provenance': 20,
        'fichier_source_html': 35,
        'plateforme_source': 18,
        'verification_requise': 12,
        'raisons_verification': 35,
        'notes_verification': 35,
        'statut': 12,
        'justification_juridique_courte': 45,
    }
    
    widths = {}
    for col in df.columns:
        if col in custom_widths:
            widths[col] = custom_widths[col]
        else:
            max_content = df[col].astype(str).str.len().max()
            header_len = len(str(col))
            widths[col] = min(max(max_content, header_len) + 2, 50)
    
    return widths


def apply_header_style(cell):
    """Applique le style d'en-tête sombre avec texte blanc."""
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin', color='34495E'),
        right=Side(style='thin', color='34495E'),
        top=Side(style='thin', color='34495E'),
        bottom=Side(style='thin', color='34495E')
    )


def apply_data_style(cell, is_alternate=False):
    """Applique le style aux cellules de données."""
    if is_alternate:
        cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    
    cell.alignment = Alignment(vertical='top', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )


def apply_conditional_formatting_statut(ws, statut_col_idx, max_row):
    """Applique la mise en forme conditionnelle basée sur le statut."""
    green_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    ws.conditional_formatting.add(
        f'{get_column_letter(statut_col_idx)}2:{get_column_letter(statut_col_idx)}{max_row}',
        CellIsRule(operator='equal', formula=['"OUVERTE"'], fill=green_fill)
    )
    
    red_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    ws.conditional_formatting.add(
        f'{get_column_letter(statut_col_idx)}2:{get_column_letter(statut_col_idx)}{max_row}',
        CellIsRule(operator='equal', formula=['"CLOTUREE"'], fill=red_fill)
    )
    
    orange_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    ws.conditional_formatting.add(
        f'{get_column_letter(statut_col_idx)}2:{get_column_letter(statut_col_idx)}{max_row}',
        CellIsRule(operator='equal', formula=['"INCONCLUSIVE"'], fill=orange_fill)
    )


def apply_procedure_color(cell, valeur):
    """Applique la couleur selon le niveau de procédure."""
    couleurs = {
        'JOUE_PROUVE': ('2F6B9A', 'FFFFFF'),
        'FORMALISEE_NEGOCIEE_DEFENSE_SECURITE': ('1E3A5F', 'FFFFFF'),
        'FORMALISEE_SANS_PREUVE_JOUE': ('E9EEF5', '374151'),
        'MAPA_SOUS_SEUIL': ('EAF3FF', '1F4E79'),
        'INDETERMINE': ('F3F4F6', '4B5563'),
    }
    
    if valeur in couleurs:
        fond, texte = couleurs[valeur]
        cell.fill = PatternFill(start_color=fond, end_color=fond, fill_type='solid')
        cell.font = Font(color=texte, bold=True)


def create_summary_sheet(wb: Workbook, df: pd.DataFrame):
    """Crée une feuille de résumé avec des statistiques."""
    ws = wb.create_sheet('Résumé', 0)
    
    ws['A1'] = 'RÉSUMÉ DES MARCHÉS PUBLICS'
    ws['A1'].font = Font(bold=True, size=16, color='2C3E50')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 30
    
    ws['A3'] = 'Généré le :'
    ws['B3'] = datetime.now().strftime('%d/%m/%Y à %H:%M')
    ws['A3'].font = Font(bold=True)
    
    row = 5
    ws[f'A{row}'] = 'STATISTIQUES GÉNÉRALES'
    ws[f'A{row}'].font = Font(bold=True, size=12, color='2C3E50')
    ws.merge_cells(f'A{row}:C{row}')
    row += 2
    
    stats = [
        ('Nombre total de marchés', len(df)),
        ('Marchés avec vérification requise', 
         df['verification_requise'].sum() if 'verification_requise' in df.columns else 'N/A'),
    ]
    
    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Distribution statuts
    row += 2
    ws[f'A{row}'] = 'DISTRIBUTION PAR STATUT'
    ws[f'A{row}'].font = Font(bold=True, size=12, color='2C3E50')
    ws.merge_cells(f'A{row}:C{row}')
    row += 2
    
    if 'statut' in df.columns:
        status_counts = df['statut'].value_counts()
        for status, count in status_counts.items():
            ws[f'A{row}'] = status
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f'{count/len(df)*100:.1f}%'
            
            if status == 'OUVERTE':
                ws[f'A{row}'].fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
            elif status == 'CLOTUREE':
                ws[f'A{row}'].fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            elif status == 'INCONCLUSIVE':
                ws[f'A{row}'].fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
            
            row += 1
    
    # Distribution famille procédure
    row += 2
    ws[f'A{row}'] = 'DISTRIBUTION PAR FAMILLE DE PROCÉDURE'
    ws[f'A{row}'].font = Font(bold=True, size=12, color='2C3E50')
    ws.merge_cells(f'A{row}:C{row}')
    row += 2
    
    if 'famille_procedure_deduite' in df.columns:
        proc_counts = df['famille_procedure_deduite'].value_counts()
        for proc, count in proc_counts.items():
            ws[f'A{row}'] = proc
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f'{count/len(df)*100:.1f}%'
            apply_procedure_color(ws[f'A{row}'], proc)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
    
    # Champs manquants
    row += 2
    ws[f'A{row}'] = 'CHAMPS MANQUANTS (COLONNES CLÉS)'
    ws[f'A{row}'].font = Font(bold=True, size=12, color='2C3E50')
    ws.merge_cells(f'A{row}:C{row}')
    row += 2
    
    key_columns = ['reference', 'titre', 'acheteur', 'date_limite_remise_offres', 'url_marche']
    for col in key_columns:
        if col in df.columns:
            missing = df[col].isna().sum() + (df[col] == '').sum()
            ws[f'A{row}'] = col
            ws[f'B{row}'] = missing
            ws[f'C{row}'] = f'{missing/len(df)*100:.1f}%' if len(df) > 0 else '0%'
            if missing > 0:
                ws[f'B{row}'].font = Font(color='E74C3C', bold=True)
            row += 1
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    
    return ws


def convert_csv_to_excel(input_path: Path, output_path: Path) -> Path:
    """
    Convertit un fichier CSV en Excel formaté.
    
    Args:
        input_path: Chemin vers le fichier CSV source
        output_path: Chemin de sortie pour le fichier Excel
    
    Returns:
        Chemin du fichier Excel généré
    """
    print(f"📖 Lecture CSV : {input_path}")
    
    df = pd.read_csv(input_path, sep=',', quotechar='"', encoding='utf-8')
    
    print(f"✅ {len(df)} lignes chargées")
    
    # Conversion des dates
    if 'date_limite_remise_offres' in df.columns:
        df['date_limite_remise_offres'] = pd.to_datetime(
            df['date_limite_remise_offres'],
            errors='coerce',
            format='mixed'
        )
    
    # Ajout colonne statut
    if 'statut' not in df.columns:
        print("🏷️ Inférence des statuts...")
        df['statut'] = df.apply(infer_statut, axis=1)
    
    # Création workbook
    wb = Workbook()
    
    # Feuille résumé
    create_summary_sheet(wb, df)
    
    # Suppression feuille par défaut
    wb.remove(wb.active)
    
    # Feuille principale
    ws = wb.create_sheet('Marchés')
    
    # Réordonner colonnes
    priority_cols = ['statut', 'reference', 'titre', 'acheteur', 'date_limite_remise_offres']
    if 'famille_procedure_deduite' in df.columns:
        priority_cols.append('famille_procedure_deduite')
    if 'typologie_marche_verifiee' in df.columns:
        priority_cols.append('typologie_marche_verifiee')
    
    other_cols = [c for c in df.columns if c not in priority_cols]
    df_ordered = df[priority_cols + other_cols]
    
    # Écriture
    for r_idx, row in enumerate(dataframe_to_rows(df_ordered, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:
                apply_header_style(cell)
            else:
                is_alternate = (r_idx % 2 == 0)
                apply_data_style(cell, is_alternate)
                
                col_name = df_ordered.columns[c_idx - 1]
                
                # Format dates
                if col_name == 'date_limite_remise_offres' and value:
                    if isinstance(value, datetime):
                        cell.number_format = 'DD/MM/YYYY HH:MM'
                
                # Couleur procédure
                if col_name == 'famille_procedure_deduite':
                    apply_procedure_color(cell, str(value))
    
    # Largeurs colonnes
    column_widths = get_column_widths(df_ordered)
    for i, col_name in enumerate(df_ordered.columns, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = column_widths.get(col_name, 15)
    
    # Hauteur lignes
    ws.row_dimensions[1].height = 35
    for i in range(2, len(df) + 2):
        ws.row_dimensions[i].height = 40
    
    # Figer première ligne
    ws.freeze_panes = 'A2'
    
    # Filtres
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df_ordered.columns))}{len(df) + 1}"
    
    # Mise en forme conditionnelle statut
    if 'statut' in df_ordered.columns:
        statut_col_idx = df_ordered.columns.get_loc('statut') + 1
        apply_conditional_formatting_statut(ws, statut_col_idx, len(df) + 1)
    
    # Sauvegarde
    wb.save(output_path)
    print(f"\n✅ Excel généré : {output_path}")
    
    # Stats
    print(f"\n📈 Résumé:")
    print(f"   - Total marchés : {len(df)}")
    if 'statut' in df.columns:
        for status, count in df['statut'].value_counts().items():
            print(f"   - {status} : {count}")
    
    return output_path


def run_excel_export(input_csv: Path, output_excel: Path) -> Dict:
    """Exécute l'export Excel complet."""
    output_path = convert_csv_to_excel(input_csv, output_excel)
    
    return {
        'output_excel': str(output_path),
        'input_csv': str(input_csv),
    }


def print_excel_summary(stats: Dict) -> None:
    """Affiche le résumé de l'export Excel."""
    print(f"\n[EXCEL_EXPORT] Export Excel terminé")
    print("=" * 50)
    print(f"  Fichier Excel: {stats.get('output_excel', 'N/A')}")
    print("=" * 50)
