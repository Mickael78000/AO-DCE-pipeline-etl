#!/usr/bin/env python3
"""
Script d'enrichissement juridique des marchés publics - VERSION DÉFENSIVE ET AUDITABLE.

Principes:
- Classification conservatrice
- Détection explicite des conflits
- Traçabilité de toutes les décisions
- Prudence en cas de doute
"""

import pandas as pd
import argparse
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


# =============================================================================
# CONSTANTES ET CONFIGURATION
# =============================================================================

SEUILS_FORMALISE = {
    'autorite_publique_centrale': 140000,
    'collectivite_territoriale_etablissement_public': 216000,
    'entite_adjudicatrice': 432000,
}

# Marqueurs pour la détection du régime défense/sécurité
MARQUEURS_DEFENSE_SECURITE = [
    "marché de défense et sécurité",
    "marche de defense et securite",
    "défense et sécurité",
    "defense et securite",
    "défense nationale",
    "defense nationale",
    "sécurité nationale",
    "securite nationale",
    "ministère des armées",
    "ministere des armees",
    "ministère de la défense",
    "ministere de la defense",
    "secret défense",
    "secret defense",
    "confidentiel défense",
    "dirisi",
    "cnd",
]

# Marqueurs de procédure négociée formalisée (PRIORITAIRES sur MAPA)
MARQUEURS_NEGOCIEE = [
    "procédure avec négociation",
    "procedure avec negociation",
    "procédure négociée",
    "procedure negociee",
    "concurrentielle avec négociation",
    "concurrentielle avec negociation",
    "négociée avec publication préalable",
    "negociee avec publication préalable",
    "négociée avec publication préalable d'un appel à la concurrence",
    "negociee avec publication prealable d'un appel a la concurrence",
    "appel d'offres avec négociation",
    "appel d offres avec négociation",
    "appel d'offres avec negociation",
]

# Marqueurs MAPA - UNIQUEMENT pour detection explicite
MARQUEURS_MAPA = [
    "mapa",
    "procédure adaptée",
    "procedure adaptee",
    "procedure adaptée",
]

# Marqueurs JOUE
MARQUEURS_JOUE = [
    "/joue/",
    "journal officiel de l'union européenne",
    "journal officiel de l'union europeenne",
]

# Couleurs pour l'Excel
COULEURS = {
    'JOUE_PROUVE': {'fond': '2F6B9A', 'texte': 'FFFFFF'},
    'FORMALISEE_NEGOCIEE_DEFENSE_SECURITE': {'fond': '1E3A5F', 'texte': 'FFFFFF'},
    'FORMALISEE_SANS_PREUVE_JOUE': {'fond': 'E9EEF5', 'texte': '374151'},
    'MAPA_SOUS_SEUIL': {'fond': 'EAF3FF', 'texte': '1F4E79'},
    'INDETERMINE': {'fond': 'F3F4F6', 'texte': '4B5563'},
}


# =============================================================================
# FONCTIONS DE NORMALISATION ET UTILITAIRES
# =============================================================================

def normaliser_texte(value) -> str:
    """
    Normalise un texte pour la comparaison:
    - minuscules
    - remplacement des caractères typographiques
    - homogénéisation des espaces
    """
    if value is None:
        return ""
    text = str(value).lower()
    # Remplacement des caractères typographiques
    text = text.replace("'", "'").replace("'", "'")
    text = text.replace("œ", "oe").replace("æ", "ae")
    text = text.replace("–", "-").replace("—", "-")
    # Normalisation des espaces
    text = " ".join(text.split())
    return text


def construire_texte_source(row: dict, champs: list) -> str:
    """
    Concatène plusieurs champs pour la recherche de marqueurs.
    Ignore les valeurs nulles.
    """
    valeurs = []
    for champ in champs:
        val = row.get(champ)
        if val is not None and str(val).strip():
            valeurs.append(normaliser_texte(val))
    return " ".join(valeurs)


def contient_marqueur(texte: str, marqueurs: list) -> tuple:
    """
    Vérifie si un texte contient l'un des marqueurs.
    Retourne: (trouvé: bool, marqueur_trouvé: str ou None)
    """
    for marqueur in marqueurs:
        if marqueur in texte:
            return True, marqueur
    return False, None


# =============================================================================
# FONCTIONS DE DÉTECTION (retournent tuples avec preuve)
# =============================================================================

def detecter_categorie_regime(row: dict) -> tuple:
    """
    Détecte le régime juridique du marché.
    
    Retourne: (categorie_regime, preuve_regime)
    
    Valeurs:
    - defense_securite: uniquement si marqueur explicite et fort
    - marches_ordinaires: par défaut
    - indetermine: si ambiguïté réelle
    """
    champs_recherche = [
        'reference', 'titre', 'notes_verification', 'raisons_verification',
        'procedure_type', 'acheteur'
    ]
    texte_complet = construire_texte_source(row, champs_recherche)
    
    trouve, marqueur = contient_marqueur(texte_complet, MARQUEURS_DEFENSE_SECURITE)
    
    if trouve:
        return "defense_securite", f"Marqueur explicite: '{marqueur}'"
    
    return "marches_ordinaires", "Aucun marqueur défense/sécurité détecté"


def detecter_typologie_marche(row: dict) -> tuple:
    """
    Détecte la typologie du marché selon sa nature réelle.
    
    Retourne: (typologie, preuve_typologie)
    
    Valeurs: fournitures, services, travaux, indetermine
    """
    type_marche = normaliser_texte(row.get('type_marche', ''))
    titre = normaliser_texte(row.get('titre', ''))
    ccag = normaliser_texte(row.get('ccag_type', ''))
    cpv = normaliser_texte(row.get('cpv_principal', ''))
    
    indices = []
    
    # Mots-clés pour travaux
    mots_travaux = ['travaux', 'construction', 'chantier', 'réhabilitation', 'rehabilitation', 
                    'gros oeuvre', 'gros-oeuvre', 'bâtiment', 'batiment', 'ouvrage']
    # Mots-clés pour fournitures
    mots_fournitures = ['fourniture', 'serveur', 'licence', 'matériel', 'materiel', 
                        'logiciel', 'équipement', 'equipement', 'hardware', 'material']
    # Mots-clés pour services
    mots_services = ['service', 'prestation', 'assistance', 'conseil', 'expertise',
                     'infogérance', 'infogerance', 'maintenance', 'développement', 
                     'developpement', 'cloud', 'hébergement', 'hebergement', 'tma',
                     'migration', 'intégration', 'integration', 'support']
    
    # Vérification via type_marche (forte)
    if type_marche == 'travaux':
        indices.append("type_marche='travaux'")
    elif type_marche == 'fournitures':
        indices.append("type_marche='fournitures'")
    elif type_marche == 'services':
        indices.append("type_marche='services'")
    
    # Vérification via CCAG
    if 'travaux' in ccag:
        indices.append("CCAG travaux")
    elif 'fourniture' in ccag:
        indices.append("CCAG fournitures")
    elif any(x in ccag for x in ['prestations_intellectuelles', 'tic', 'prestation']):
        indices.append("CCAG prestations/services")
    
    # Vérification via titre
    titre_lower = titre
    for mot in mots_travaux:
        if mot in titre_lower:
            indices.append(f"titre contient '{mot}'")
            break
    for mot in mots_fournitures:
        if mot in titre_lower:
            indices.append(f"titre contient '{mot}'")
            break
    for mot in mots_services:
        if mot in titre_lower:
            indices.append(f"titre contient '{mot}'")
            break
    
    # Décompte par catégorie
    count_travaux = sum(1 for m in mots_travaux if m in titre_lower)
    count_fournitures = sum(1 for m in mots_fournitures if m in titre_lower)
    count_services = sum(1 for m in mots_services if m in titre_lower)
    
    # Décision avec règle de majorité
    max_count = max(count_travaux, count_fournitures, count_services)
    
    if count_travaux > 0 and count_travaux == max_count:
        return "travaux", "; ".join(indices) if indices else "indice titre"
    elif count_fournitures > 0 and count_fournitures == max_count:
        return "fournitures", "; ".join(indices) if indices else "indice titre"
    elif count_services > 0 and count_services == max_count:
        return "services", "; ".join(indices) if indices else "indice titre"
    
    # Si type_marche défini mais pas d'indice titre
    if type_marche in ['travaux', 'fournitures', 'services']:
        return type_marche, f"type_marche='{type_marche}' (pas d'indice titre)"
    
    return "indetermine", "Aucun indice typologique clair"


def detecter_qualite_preuve_procedure(row: dict, typologie: str, categorie_regime: str) -> tuple:
    """
    Évalue la qualité de la preuve de procédure.
    
    Retourne: (qualite_preuve, type_conflit)
    
    Qualités: EXPLICITE, FORT_INDICE, FAIBLE_INDICE, CONTRADICTOIRE, ABSENTE
    """
    procedure_type = normaliser_texte(row.get('procedure_type', ''))
    titre = normaliser_texte(row.get('titre', ''))
    notes = normaliser_texte(row.get('notes_verification', ''))
    raisons = normaliser_texte(row.get('raisons_verification', ''))
    
    # Vérification EXPLICITE dans procedure_type
    if procedure_type:
        # MAPA explicite
        trouve_mapa, _ = contient_marqueur(procedure_type, MARQUEURS_MAPA)
        # Négociée explicite
        trouve_neg, _ = contient_marqueur(procedure_type, MARQUEURS_NEGOCIEE)
        # Formalisée explicite
        trouve_form = procedure_type == 'formalisee' or 'formalisée' in procedure_type
        # JOUE
        trouve_joue = '/joue/' in procedure_type or 'joue' in procedure_type
        
        if trouve_mapa or trouve_neg or trouve_form or trouve_joue:
            return "EXPLICITE", ""
    
    # Recherche dans d'autres champs pour détecter conflits
    texte_secondaire = f"{titre} {notes} {raisons}"
    
    negociée_secondaire, _ = contient_marqueur(texte_secondaire, MARQUEURS_NEGOCIEE)
    mapa_secondaire, _ = contient_marqueur(texte_secondaire, MARQUEURS_MAPA)
    
    # Détection des conflits
    if procedure_type:
        # Conflit: MAPA explicite mais négociation dans autres champs
        if trouve_mapa and negociée_secondaire:
            return "CONTRADICTOIRE", "MAPA explicite mais mention de négociation dans autres champs"
        
        # Conflit: Formalisée explicite mais MAPA dans autres champs
        if trouve_form and mapa_secondaire:
            return "CONTRADICTOIRE", "Formalisée explicite mais mention MAPA dans autres champs"
        
        # Conflit: Défense/sécurité + MAPA
        if categorie_regime == "defense_securite" and trouve_mapa:
            return "CONTRADICTOIRE", "Régime défense/sécurité avec MAPA explicite"
    
    # FORT_INDICE: plusieurs champs convergent
    if negociée_secondaire:
        return "FORT_INDICE", ""
    
    # FAIBLE_INDICE: un seul champ suggère
    if 'procedure' in texte_secondaire or 'marche' in texte_secondaire:
        return "FAIBLE_INDICE", ""
    
    return "ABSENTE", ""


def est_mapa_explicite(row: dict) -> bool:
    """
    Vérifie si procedure_type contient explicitement MAPA.
    UNIQUEMENT dans procedure_type.
    """
    procedure_type = normaliser_texte(row.get('procedure_type', ''))
    trouve, _ = contient_marqueur(procedure_type, MARQUEURS_MAPA)
    return trouve


def est_procedure_negociee(row: dict) -> tuple:
    """
    Vérifie si une procédure négociée est explicitement mentionnée.
    Cherche dans plusieurs champs.
    
    Retourne: (est_negociee: bool, marqueur_trouvé: str)
    """
    champs = ['procedure_type', 'titre', 'notes_verification', 'raisons_verification']
    texte = construire_texte_source(row, champs)
    return contient_marqueur(texte, MARQUEURS_NEGOCIEE)


def est_joue_prove(row: dict) -> tuple:
    """
    Vérifie si une preuve JOUE est présente.
    
    Retourne: (est_prouvé: bool, source_preuve: str)
    """
    reference = normaliser_texte(row.get('reference', ''))
    url_marche = normaliser_texte(row.get('url_marche', ''))
    fichier_source = normaliser_texte(row.get('fichier_source_html', ''))
    
    # Référence /joue/
    if '/joue/' in reference:
        return True, 'reference'
    
    # URL /joue/
    if '/joue/' in url_marche:
        return True, 'url'
    
    # Fichier HTML
    if 'joue' in fichier_source:
        return True, 'html'
    
    # Texte
    champs = ['notes_verification', 'raisons_verification', 'titre']
    texte = construire_texte_source(row, champs)
    trouve, _ = contient_marqueur(texte, MARQUEURS_JOUE)
    if trouve:
        return True, 'texte_source'
    
    return False, ''


# =============================================================================
# FONCTION PRINCIPALE DE DÉDUCTION
# =============================================================================

def deduire_niveau_procedure(row: dict, categorie_regime: str, qualite_preuve: str, 
                                type_conflit: str) -> tuple:
    """
    Déduit le niveau de procédure selon l'arbre de décision strict.
    
    Retourne: (niveau_procedure, justification)
    """
    # Étape 1: Gestion des conflits
    if type_conflit and qualite_preuve == "CONTRADICTOIRE":
        # Conflit majeur: prudence
        if "défense/sécurité avec MAPA" in type_conflit:
            return "INDETERMINE", f"Conflit majeur: {type_conflit}. Régime défense/sécurité incompatible avec MAPA."
        if "MAPA explicite mais mention de négociation" in type_conflit:
            # Priorité à la négociation (formalisée) sur MAPA
            return "FORMALISEE_SANS_PREUVE_JOUE", f"Conflit résolu: procédure avec négociation prime sur MAPA."
        return "INDETERMINE", f"Conflit détecté: {type_conflit}. Classification prudente."
    
    # Étape 2: Régime défense/sécurité
    if categorie_regime == "defense_securite":
        neg_trouve, marqueur_neg = est_procedure_negociee(row)
        if neg_trouve:
            return "FORMALISEE_NEGOCIEE_DEFENSE_SECURITE", \
                   f"Régime défense/sécurité ; procédure négociée formalisée détectée ('{marqueur_neg}')."
        
        joue_trouve, _ = est_joue_prove(row)
        if joue_trouve:
            return "JOUE_PROUVE", "Régime défense/sécurité ; publication JOUE avérée."
        
        return "INDETERMINE", "Régime défense/sécurité ; procédure insuffisamment décrite."
    
    # Étape 3: Régime ordinaire
    
    # 3.1 Procédure négociée (PRIORITAIRE sur MAPA)
    neg_trouve, marqueur_neg = est_procedure_negociee(row)
    if neg_trouve:
        return "FORMALISEE_SANS_PREUVE_JOUE", \
               f"Procédure avec négociation détectée ('{marqueur_neg}') ; classification formalisée."
    
    # 3.2 JOUE prouvée
    joue_trouve, _ = est_joue_prove(row)
    if joue_trouve:
        return "JOUE_PROUVE", "Publication JOUE avérée."
    
    # 3.3 MAPA explicite (UNIQUEMENT si pas de négociation détectée avant)
    if est_mapa_explicite(row):
        return "MAPA_SOUS_SEUIL", "Procédure explicitement qualifiée MAPA dans procedure_type."
    
    # 3.4 Formalisée explicite
    procedure_type = normaliser_texte(row.get('procedure_type', ''))
    if procedure_type in ['formalisee', 'formalisée'] or 'formalisee' in procedure_type:
        return "FORMALISEE_SANS_PREUVE_JOUE", "Procédure formalisée explicitement qualifiée."
    
    # 3.5 Indéterminé
    if qualite_preuve == "ABSENTE":
        return "INDETERMINE", "Aucune preuve de procédure détectée."
    
    return "INDETERMINE", "Preuve insuffisante pour qualifier le niveau de procédure."


# =============================================================================
# FONCTION D'ENRICHISSEMENT PRINCIPALE
# =============================================================================

def enrichir_ligne(row):
    """
    Enrichit une ligne avec toutes les colonnes juridiques.
    Architecture défensive et auditable.
    """
    result = {}
    
    # ÉTAPE 1: Détection du régime
    categorie_regime, preuve_regime = detecter_categorie_regime(row)
    result['categorie_regime'] = categorie_regime
    result['preuve_regime'] = preuve_regime
    
    # ÉTAPE 2: Détection de la typologie
    typologie, preuve_typologie = detecter_typologie_marche(row)
    result['typologie_marche_verifiee'] = typologie
    result['preuve_typologie'] = preuve_typologie
    
    # ÉTAPE 3: Qualité de preuve et conflits
    qualite_preuve, type_conflit = detecter_qualite_preuve_procedure(
        row, typologie, categorie_regime
    )
    result['qualite_preuve_procedure'] = qualite_preuve
    result['conflit_detecte'] = 'oui' if type_conflit else 'non'
    result['type_conflit'] = type_conflit
    
    # ÉTAPE 4: Préparation des preuves JOUE
    joue_trouve, source_joue = est_joue_prove(row)
    result['preuve_joue_detectee'] = 'oui' if joue_trouve else 'non'
    result['source_preuve_joue'] = source_joue
    
    # ÉTAPE 5: Déduction du niveau de procédure
    niveau, justification = deduire_niveau_procedure(
        row, categorie_regime, qualite_preuve, type_conflit
    )
    # Alias pour lisibilité : famille_procedure_deduite = niveau_procedure_deduit
    result['famille_procedure_deduite'] = niveau
    result['niveau_procedure_deduit'] = niveau  # conservé pour compatibilité
    result['justification_juridique_courte'] = justification
    
    # ÉTAPE 6: Calcul du seuil (si applicable)
    if categorie_regime == 'defense_securite' or typologie in ['travaux', 'indetermine']:
        result['seuil_formalise_applicable'] = ''
    else:
        # Calcul du seuil pour marchés ordinaires fournitures/services
        categorie_acheteur = determiner_categorie_acheteur(row)
        seuil = determiner_seuil_applicable(categorie_acheteur, categorie_regime)
        result['seuil_formalise_applicable'] = seuil if seuil else ''
    
    # ÉTAPE 7: Code couleur
    result['code_couleur_procedure'] = niveau
    
    return result


def determiner_categorie_acheteur(row):
    """Détermine la catégorie d'acheteur pour les seuils."""
    type_acheteur = normaliser_texte(row.get('type_acheteur', ''))
    fonction_publique = normaliser_texte(row.get('fonction_publique', ''))
    acheteur = normaliser_texte(row.get('acheteur', ''))
    
    if 'etat' in type_acheteur or 'etat' in fonction_publique or \
       any(term in acheteur for term in ['ministere', 'ministère', 'dgfip']):
        return 'autorite_publique_centrale'
    
    if any(term in type_acheteur for term in ['collectivite', 'etablissement_public', 'groupement']):
        return 'collectivite_territoriale_etablissement_public_autre_acheteur'
    
    if 'entite_adjudicatrice' in type_acheteur or 'adjudicatrice' in type_acheteur:
        return 'entite_adjudicatrice'
    
    return 'indetermine'


def determiner_seuil_applicable(categorie_acheteur, categorie_regime):
    """Détermine le seuil applicable."""
    if categorie_regime == 'defense_securite':
        return None
    
    if categorie_acheteur == 'autorite_publique_centrale':
        return SEUILS_FORMALISE['autorite_publique_centrale']
    elif categorie_acheteur == 'collectivite_territoriale_etablissement_public_autre_acheteur':
        return SEUILS_FORMALISE['collectivite_territoriale_etablissement_public']
    elif categorie_acheteur == 'entite_adjudicatrice':
        return SEUILS_FORMALISE['entite_adjudicatrice']
    
    return None


# =============================================================================
# FONCTIONS D'ENRICHISSEMENT DATAFRAME ET EXPORT
# =============================================================================

def enrichir_dataframe(df):
    """Enrichit tout le dataframe avec ordre optimisé des colonnes."""
    print("🔍 Analyse juridique défensive des marchés en cours...")
    
    enrichments = []
    for _, row in df.iterrows():
        enrichments.append(enrichir_ligne(row))
    
    df_enrich = pd.DataFrame(enrichments)
    df_result = pd.concat([df, df_enrich], axis=1)
    
    # Réorganiser les colonnes selon l'ordre optimisé
    ordered_cols = [
        # BLOC 1: Identification & Décision
        'reference', 'titre', 'acheteur', 'type_acheteur', 'fonction_publique',
        'procedure_type', 'famille_procedure_deduite', 'typologie_marche_verifiee',
        'seuil_formalise_applicable', 'montant_estime', 'date_limite_remise_offres',
        # BLOC 2: Sources
        'url_marche', 'fichier_source_html', 'plateforme_source',
        # BLOC 3: Contexte
        'duree', 'renouvellements', 'localisation', 'cpv_principal', 'cpv_secondaires', 'ccag_type',
        # BLOC 4: Audit
        'verification_requise', 'raisons_verification', 'notes_verification',
        # BLOC 5: Preuves
        'qualite_preuve_procedure', 'conflit_detecte', 'type_conflit',
        'categorie_regime', 'preuve_regime', 'preuve_typologie',
        'preuve_joue_detectee', 'source_preuve_joue',
        'justification_juridique_courte', 'code_couleur_procedure', 'niveau_procedure_deduit',
        # BLOC 6: Provenance
        'url_provenance',
    ]
    
    # Garder uniquement les colonnes qui existent
    final_cols = [c for c in ordered_cols if c in df_result.columns]
    # Ajouter les colonnes restantes non listées
    remaining = [c for c in df_result.columns if c not in ordered_cols]
    final_cols.extend(remaining)
    
    df_result = df_result[final_cols]
    
    # Statistiques
    stats = df_enrich['famille_procedure_deduite'].value_counts()
    print(f"\n📊 Résultats de la qualification:")
    for niveau, count in stats.items():
        print(f"   - {niveau}: {count}")
    
    # Conflits détectés
    conflits = df_enrich['conflit_detecte'].value_counts()
    if 'oui' in conflits:
        print(f"\n⚠️  Conflits détectés: {conflits['oui']} marchés")
    
    return df_result


def create_methode_sheet(wb):
    """Crée la feuille méthodologie."""
    ws = wb.create_sheet('Méthode', 0)
    
    ws['A1'] = 'MÉTHODOLOGIE DE CLASSIFICATION DÉFENSIVE'
    ws['A1'].font = Font(bold=True, size=14, color='2C3E50')
    ws.merge_cells('A1:E1')
    
    sections = [
        ('', []),
        ('PRINCIPES DIRECTEURS', [
            '• Classification conservatrice et prudente',
            '• Détection explicite des conflits de source',
            '• Traçabilité de toutes les décisions',
            '• Préférence pour INDETERMINE en cas de doute',
        ]),
        ('', []),
        ('ORDRE DE DÉCISION STRICT', [
            '1. Détection des conflits entre champs',
            '2. Détection du régime (défense/sécurité vs ordinaire)',
            '3. Détection de la typologie (fournitures/services/travaux)',
            '4. Évaluation de la qualité de preuve',
            '5. Déduction du niveau de procédure',
        ]),
        ('', []),
        ('RÈGLES IMPÉRATIVES', [
            '• Procédure avec négociation = formalisée (JAMAIS MAPA)',
            '• MAPA uniquement si explicite dans procedure_type',
            '• Défense/sécurité explicite requise (pas d\'inférence contextuelle)',
            '• Typologie indépendante du régime',
            '• Conflit = prudence ou INDETERMINE',
        ]),
        ('', []),
        ('QUALITÉS DE PREUVE', [
            '• EXPLICITE: procédure_type contient la procédure',
            '• FORT_INDICE: plusieurs champs convergent',
            '• FAIBLE_INDICE: un seul champ suggère',
            '• CONTRADICTOIRE: champs opposés',
            '• ABSENTE: aucune preuve utile',
        ]),
        ('', []),
        ('ORDRE DES COLONNES (optimisé pour revue manuelle)', [
            'BLOC 1 - Identification & Décision:',
            '  reference, titre, acheteur, type_acheteur, fonction_publique,',
            '  procedure_type, famille_procedure_deduite, typologie_marche_verifiee,',
            '  seuil_formalise_applicable, montant_estime, date_limite_remise_offres',
            '',
            'BLOC 2 - Sources vérifiables:',
            '  url_marche, fichier_source_html, plateforme_source',
            '',
            'BLOC 3 - Contexte métier:',
            '  duree, renouvellements, localisation, cpv_principal,',
            '  cpv_secondaires, ccag_type',
            '',
            'BLOC 4 - Audit:',
            '  verification_requise, raisons_verification, notes_verification',
            '',
            'BLOC 5 - Preuves et qualité:',
            '  qualite_preuve_procedure, conflit_detecte, type_conflit,',
            '  categorie_regime, preuve_regime, preuve_typologie,',
            '  preuve_joue_detectee, source_preuve_joue,',
            '  justification_juridique_courte, code_couleur_procedure',
            '',
            'BLOC 6 - Provenance:',
            '  url_provenance',
        ]),
    ]
    
    row_num = 3
    for titre, items in sections:
        if titre:
            ws.cell(row=row_num, column=1, value=titre).font = Font(bold=True, size=12, color='2C3E50')
            ws.merge_cells(f'A{row_num}:E{row_num}')
            row_num += 1
        for item in items:
            ws.cell(row=row_num, column=1, value=item)
            row_num += 1
    
    ws.column_dimensions['A'].width = 80
    return ws


def create_excel_formatted(df, output_path):
    """Crée le fichier Excel formaté avec ordre optimisé pour revue manuelle."""
    wb = Workbook()
    create_methode_sheet(wb)
    wb.remove(wb.active)
    
    ws = wb.create_sheet('Marchés Qualifiés')
    
    # Ordre des colonnes optimisé pour la revue manuelle
    # BLOC 1: Identification et Décision (prioritaire pour lecture)
    bloc_identification = [
        'reference',
        'titre',
        'acheteur',
        'type_acheteur',
        'fonction_publique',
        'procedure_type',
        'famille_procedure_deduite',
        'typologie_marche_verifiee',
        'seuil_formalise_applicable',
        'montant_estime',
        'date_limite_remise_offres',
    ]
    
    # BLOC 2: Sources vérifiables (URLs visibles tôt)
    bloc_sources = [
        'url_marche',
        'fichier_source_html',
        'plateforme_source',
    ]
    
    # BLOC 3: Contexte métier
    bloc_contexte = [
        'duree',
        'renouvellements',
        'localisation',
        'cpv_principal',
        'cpv_secondaires',
        'ccag_type',
    ]
    
    # BLOC 4: Audit et Vérification
    bloc_audit = [
        'verification_requise',
        'raisons_verification',
        'notes_verification',
    ]
    
    # BLOC 5: Preuves et Qualité (indicateurs de confiance)
    bloc_preuves = [
        'qualite_preuve_procedure',
        'conflit_detecte',
        'type_conflit',
        'categorie_regime',
        'preuve_regime',
        'preuve_typologie',
        'preuve_joue_detectee',
        'source_preuve_joue',
        'justification_juridique_courte',
        'code_couleur_procedure',
        'niveau_procedure_deduit',  # conservé pour compatibilité
    ]
    
    # BLOC 6: Provenance technique (en dernier)
    bloc_provenance = [
        'url_provenance',
    ]
    
    # Colonnes dans l'ordre souhaité
    ordered_cols = (
        bloc_identification +
        bloc_sources +
        bloc_contexte +
        bloc_audit +
        bloc_preuves +
        bloc_provenance
    )
    
    # Ajouter les colonnes restantes qui ne sont pas dans l'ordre défini
    remaining_cols = [c for c in df.columns if c not in ordered_cols]
    final_cols = [c for c in ordered_cols if c in df.columns] + remaining_cols
    
    df_ordered = df[final_cols]
    
    # Écriture
    for r_idx, row in enumerate(dataframe_to_rows(df_ordered, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:
                cell.font = Font(bold=True, color='FFFFFF', size=10)
                cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if r_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
                
                col_name = df_ordered.columns[c_idx - 1]
                # Coloration pour famille_procedure_deduite
                if col_name == 'famille_procedure_deduite':
                    couleurs = COULEURS.get(str(value), COULEURS['INDETERMINE'])
                    cell.fill = PatternFill(start_color=couleurs['fond'], end_color=couleurs['fond'], fill_type='solid')
                    cell.font = Font(color=couleurs['texte'], bold=True)
    
    # Ajustement largeurs par bloc
    col_idx = 1
    # Bloc 1: Identification (colonnes larges pour lisibilité)
    for _ in bloc_identification[:2]:  # reference, titre
        ws.column_dimensions[get_column_letter(col_idx)].width = 35
        col_idx += 1
    for _ in bloc_identification[2:6]:  # acheteur, type_acheteur, fonction_publique, procedure_type
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
        col_idx += 1
    for _ in bloc_identification[6:]:  # famille_procedure, typologie, seuil, montant, date
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
        col_idx += 1
    
    # Bloc 2: Sources (URLs)
    for _ in bloc_sources:
        ws.column_dimensions[get_column_letter(col_idx)].width = 40
        col_idx += 1
    
    # Bloc 3: Contexte
    for _ in bloc_contexte:
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
        col_idx += 1
    
    # Bloc 4: Audit
    for _ in bloc_audit:
        ws.column_dimensions[get_column_letter(col_idx)].width = 30
        col_idx += 1
    
    # Bloc 5: Preuves
    for col in bloc_preuves:
        if col in df.columns:
            if col == 'justification_juridique_courte':
                ws.column_dimensions[get_column_letter(col_idx)].width = 50
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            col_idx += 1
    
    # Bloc 6: Provenance
    ws.column_dimensions[get_column_letter(col_idx)].width = 40
    
    # Gel du volet à la colonne G (après famille_procedure_deduite)
    ws.freeze_panes = 'G2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df_ordered.columns))}{len(df) + 1}"
    
    wb.save(output_path)
    print(f"\n✅ Fichier Excel qualifié généré: {output_path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Enrichit un fichier de marchés publics avec qualification juridique défensive',
    )
    parser.add_argument('input', help='Chemin vers le fichier CSV source')
    parser.add_argument('-o', '--output-csv', help='Chemin du CSV de sortie')
    parser.add_argument('-e', '--excel', help='Chemin du Excel de sortie')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.input):
        print(f"❌ Erreur: Le fichier {args.input} n'existe pas.")
        return 1
    
    try:
        print(f"📖 Lecture: {args.input}")
        df = pd.read_csv(args.input, sep=',', quotechar='"', encoding='utf-8')
        print(f"✅ {len(df)} lignes chargées")
        
        df_enriched = enrichir_dataframe(df)
        
        # Sauvegarde CSV
        csv_path = args.output_csv or f"{os.path.splitext(args.input)[0]}-juridique-v8.csv"
        df_enriched.to_csv(csv_path, index=False, encoding='utf-8')
        print(f"✅ CSV sauvegardé: {csv_path}")
        
        # Excel
        if args.excel:
            create_excel_formatted(df_enriched, args.excel)
        
        print(f"\n🎉 Enrichissement terminé avec succès!")
        return 0
        
    except Exception as e:
        print(f"❌ Erreur: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    exit(main())
