#!/usr/bin/env python3
"""
Script de conversion CSV vers Excel formaté pour les marchés publics.
Génère un fichier Excel lisible avec mise en forme professionnelle.
"""

import pandas as pd
import argparse
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule


def infer_statut(row):
    """
    Déduit le statut du marché à partir de la date limite.
    """
    date_limite = row.get('date_limite_remise_offres')
    if pd.isna(date_limite) or date_limite == '':
        return 'INCONCLUSIVE'
    
    try:
        if isinstance(date_limite, str):
            date_limite = pd.to_datetime(date_limite)
        
        if date_limite < datetime.now():
            return 'CLOTUREE'
        else:
            return 'OUVERTE'
    except:
        return 'INCONCLUSIVE'


def get_column_widths(df):
    """
    Définit les largeurs de colonnes personnalisées.
    """
    # Largeurs par défaut pour certaines colonnes
    custom_widths = {
        'reference': 18,
        'titre': 50,
        'acheteur': 35,
        'type_acheteur': 18,
        'fonction_publique': 18,
        'procedure_type': 15,
        'type_marche': 12,
        'ccag_type': 22,
        'cpv_principal': 15,
        'cpv_secondaires': 35,
        'localisation': 25,
        'date_limite_remise_offres': 22,
        'duree': 10,
        'renouvellements': 30,
        'montant_estime': 18,
        'url_marche': 50,
        'url_provenance': 20,
        'fichier_source_html': 35,
        'plateforme_source': 18,
        'verification_requise': 12,
        'raisons_verification': 35,
        'notes_verification': 35,
        'statut': 12,
    }
    
    widths = {}
    for col in df.columns:
        if col in custom_widths:
            widths[col] = custom_widths[col]
        else:
            # Calculer la largeur basée sur le contenu
            max_content = df[col].astype(str).str.len().max()
            header_len = len(str(col))
            widths[col] = min(max(max_content, header_len) + 2, 50)
    
    return widths


def apply_header_style(cell):
    """Applique le style d'en-tête sombre avec texte blanc."""
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin', color='34495E'),
        right=Side(style='thin', color='34495E'),
        top=Side(style='thin', color='34495E'),
        bottom=Side(style='thin', color='34495E')
    )


def apply_data_style(cell, is_alternate=False):
    """Applique le style aux cellules de données."""
    if is_alternate:
        cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    
    cell.alignment = Alignment(vertical='top', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )


def apply_conditional_formatting(ws, statut_col_idx, max_row):
    """Applique la mise en forme conditionnelle basée sur le statut."""
    # Vert pour OUVERTE
    green_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    ws.conditional_formatting.add(
        f'{get_column_letter(statut_col_idx)}2:{get_column_letter(statut_col_idx)}{max_row}',
        CellIsRule(operator='equal', formula=['"OUVERTE"'], fill=green_fill)
    )
    
    # Rouge/Gris pour CLOTUREE
    red_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    ws.conditional_formatting.add(
        f'{get_column_letter(statut_col_idx)}2:{get_column_letter(statut_col_idx)}{max_row}',
        CellIsRule(operator='equal', formula=['"CLOTUREE"'], fill=red_fill)
    )
    
    # Orange pour INCONCLUSIVE
    orange_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    ws.conditional_formatting.add(
        f'{get_column_letter(statut_col_idx)}2:{get_column_letter(statut_col_idx)}{max_row}',
        CellIsRule(operator='equal', formula=['"INCONCLUSIVE"'], fill=orange_fill)
    )


def create_summary_sheet(wb, df):
    """Crée une feuille de résumé avec des statistiques."""
    ws = wb.create_sheet('Résumé', 0)
    
    # Titre
    ws['A1'] = 'RÉSUMÉ DES MARCHÉS PUBLICS'
    ws['A1'].font = Font(bold=True, size=16, color='2C3E50')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 30
    
    # Date de génération
    ws['A3'] = 'Généré le :'
    ws['B3'] = datetime.now().strftime('%d/%m/%Y à %H:%M')
    ws['A3'].font = Font(bold=True)
    
    # Statistiques générales
    row = 5
    ws[f'A{row}'] = 'STATISTIQUES GÉNÉRALES'
    ws[f'A{row}'].font = Font(bold=True, size=12, color='2C3E50')
    ws.merge_cells(f'A{row}:C{row}')
    row += 2
    
    stats = [
        ('Nombre total de marchés', len(df)),
        ('Marchés avec vérification requise', df['verification_requise'].sum() if 'verification_requise' in df.columns else 'N/A'),
    ]
    
    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Distribution des statuts
    row += 2
    ws[f'A{row}'] = 'DISTRIBUTION PAR STATUT'
    ws[f'A{row}'].font = Font(bold=True, size=12, color='2C3E50')
    ws.merge_cells(f'A{row}:C{row}')
    row += 2
    
    if 'statut' in df.columns:
        status_counts = df['statut'].value_counts()
        for status, count in status_counts.items():
            ws[f'A{row}'] = status
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f'{count/len(df)*100:.1f}%'
            
            # Couleur selon le statut
            if status == 'OUVERTE':
                ws[f'A{row}'].fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
            elif status == 'CLOTUREE':
                ws[f'A{row}'].fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            elif status == 'INCONCLUSIVE':
                ws[f'A{row}'].fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
            
            row += 1
    
    # Distribution par type d'acheteur
    row += 2
    ws[f'A{row}'] = 'DISTRIBUTION PAR TYPE ACHETEUR'
    ws[f'A{row}'].font = Font(bold=True, size=12, color='2C3E50')
    ws.merge_cells(f'A{row}:C{row}')
    row += 2
    
    if 'type_acheteur' in df.columns:
        buyer_counts = df['type_acheteur'].value_counts().head(10)
        for buyer_type, count in buyer_counts.items():
            ws[f'A{row}'] = buyer_type
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f'{count/len(df)*100:.1f}%'
            row += 1
    
    # Champs manquants
    row += 2
    ws[f'A{row}'] = 'CHAMPS MANQUANTS (COLONNES CLÉS)'
    ws[f'A{row}'].font = Font(bold=True, size=12, color='2C3E50')
    ws.merge_cells(f'A{row}:C{row}')
    row += 2
    
    key_columns = ['reference', 'titre', 'acheteur', 'date_limite_remise_offres', 'url_marche']
    for col in key_columns:
        if col in df.columns:
            missing = df[col].isna().sum() + (df[col] == '').sum()
            ws[f'A{row}'] = col
            ws[f'B{row}'] = missing
            ws[f'C{row}'] = f'{missing/len(df)*100:.1f}%' if len(df) > 0 else '0%'
            if missing > 0:
                ws[f'B{row}'].font = Font(color='E74C3C', bold=True)
            row += 1
    
    # Ajustement des largeurs
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    
    return ws


def create_legend_sheet(wb):
    """Crée une feuille de légende expliquant les colonnes."""
    ws = wb.create_sheet('Légende')
    
    ws['A1'] = 'LÉGENDE ET DESCRIPTION DES COLONNES'
    ws['A1'].font = Font(bold=True, size=14, color='2C3E50')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 25
    
    legend_data = [
        ('Colonne', 'Description', 'Type', 'Exemple'),
        ('reference', 'Identifiant unique du marché', 'Texte', '13/joue/002671162026'),
        ('titre', 'Intitulé du marché', 'Texte long', 'Mise à disposition et gestion du workplace IT'),
        ('acheteur', 'Organisation acheteuse', 'Texte', 'Ville de Paris'),
        ('type_acheteur', 'Catégorie administrative', 'Catégorie', 'etat, collectivite_territoriale...'),
        ('fonction_publique', 'Type de fonction publique', 'Catégorie', 'etat, territoriale, hospitaliere...'),
        ('procedure_type', 'Type de procédure', 'Catégorie', 'MAPA, formalisee...'),
        ('type_marche', 'Nature du marché', 'Catégorie', 'services, fournitures...'),
        ('ccag_type', 'Clauses contractuelles applicables', 'Catégorie', 'tic, prestations_intellectuelles...'),
        ('cpv_principal', 'Code CPV principal', 'Code', '72600000'),
        ('cpv_secondaires', 'Codes CPV additionnels', 'Liste', '72600000|72500000...'),
        ('localisation', 'Zone géographique', 'Texte', 'Paris (75)'),
        ('date_limite_remise_offres', 'Date et heure limite', 'Date', '2026-05-20 12:00:00'),
        ('duree', 'Durée du contrat (mois)', 'Nombre', '48'),
        ('renouvellements', 'Modalités de reconduction', 'Texte', '3 reconduction(s) de 12 mois'),
        ('montant_estime', 'Budget estimé', 'Monétaire', '400000 EUR'),
        ('url_marche', 'Lien vers l\'annonce', 'URL', 'https://www.francemarches.com/...'),
        ('plateforme_source', 'Source de l\'annonce', 'Catégorie', 'FRANCE_MARCHES, PLACE_NUMERIC...'),
        ('verification_requise', 'Nécessite une vérification', 'Booléen', 'true/false'),
        ('raisons_verification', 'Justification si vérification requise', 'Texte', 'ccag_type_manquant...'),
        ('statut', 'État du marché (déduit)', 'Catégorie', 'OUVERTE, CLOTUREE, INCONCLUSIVE'),
    ]
    
    for i, row_data in enumerate(legend_data, start=3):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i == 3:  # Header
                apply_header_style(cell)
            else:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if i % 2 == 0:
                    cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    
    # Ajustement des largeurs
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    
    # Explications des statuts
    ws['A27'] = 'SIGNIFICATION DES STATUTS'
    ws['A27'].font = Font(bold=True, size=12, color='2C3E50')
    ws.merge_cells('A27:D27')
    
    status_explanations = [
        ('OUVERTE', 'Marché en cours, date limite non dépassée', 'D4EDDA', 'Vert'),
        ('CLOTUREE', 'Marché terminé, date limite dépassée', 'F8D7DA', 'Rouge'),
        ('INCONCLUSIVE', 'Statut indéterminé (date manquante)', 'FFF3CD', 'Orange'),
    ]
    
    for i, (status, desc, color, label) in enumerate(status_explanations, start=29):
        ws.cell(row=i, column=1, value=status).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=2, value=desc)
        ws.cell(row=i, column=3, value=label)
    
    return ws


def convert_csv_to_excel(input_path, output_path=None):
    """
    Convertit un fichier CSV en Excel formaté.
    
    Args:
        input_path: Chemin vers le fichier CSV source
        output_path: Chemin de sortie pour le fichier Excel (optionnel)
    
    Returns:
        Chemin du fichier Excel généré
    """
    # Déterminer le chemin de sortie par défaut
    if output_path is None:
        base_name = os.path.splitext(input_path)[0]
        output_path = f"{base_name}-readable.xlsx"
    
    print(f"📖 Lecture du CSV : {input_path}")
    
    # Lecture du CSV
    df = pd.read_csv(input_path, sep=',', quotechar='"', encoding='utf-8')
    
    print(f"✅ {len(df)} lignes chargées")
    print(f"📊 Colonnes : {list(df.columns)}")
    
    # Conversion des dates
    if 'date_limite_remise_offres' in df.columns:
        df['date_limite_remise_offres'] = pd.to_datetime(
            df['date_limite_remise_offres'], 
            errors='coerce',
            format='mixed'
        )
    
    # Ajout de la colonne statut si pas présente
    if 'statut' not in df.columns:
        print("🏷️  Inférence des statuts à partir des dates limites...")
        df['statut'] = df.apply(infer_statut, axis=1)
    
    # Création du workbook
    wb = Workbook()
    
    # Création de la feuille de résumé
    create_summary_sheet(wb, df)
    
    # Création de la feuille de légende
    create_legend_sheet(wb)
    
    # Suppression de la feuille par défaut
    wb.remove(wb.active)
    
    # Création de la feuille principale avec les données
    ws = wb.create_sheet('Marchés')
    
    # Réordonner les colonnes pour mettre les plus importantes en premier
    priority_cols = ['statut', 'reference', 'titre', 'acheteur', 'date_limite_remise_offres']
    other_cols = [c for c in df.columns if c not in priority_cols]
    df_ordered = df[priority_cols + other_cols]
    
    # Écriture des données
    for r_idx, row in enumerate(dataframe_to_rows(df_ordered, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:
                # Style d'en-tête
                apply_header_style(cell)
            else:
                # Style des données
                is_alternate = (r_idx % 2 == 0)
                apply_data_style(cell, is_alternate)
                
                # Formatage spécial pour les dates
                if df_ordered.columns[c_idx - 1] == 'date_limite_remise_offres' and value:
                    if isinstance(value, datetime):
                        cell.number_format = 'DD/MM/YYYY HH:MM'
    
    # Ajuster les largeurs de colonnes
    column_widths = get_column_widths(df_ordered)
    for i, col_name in enumerate(df_ordered.columns, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = column_widths.get(col_name, 15)
    
    # Ajuster la hauteur des lignes
    ws.row_dimensions[1].height = 35  # En-tête
    for i in range(2, len(df) + 2):
        ws.row_dimensions[i].height = 40  # Données avec espace pour le wrap
    
    # Figer la première ligne
    ws.freeze_panes = 'A2'
    
    # Ajouter les filtres
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df_ordered.columns))}{len(df) + 1}"
    
    # Mise en forme conditionnelle sur la colonne statut
    if 'statut' in df_ordered.columns:
        statut_col_idx = df_ordered.columns.get_loc('statut') + 1
        apply_conditional_formatting(ws, statut_col_idx, len(df) + 1)
    
    # Mise en évidence des colonnes importantes (bordure subtile)
    important_cols = ['reference', 'titre', 'acheteur', 'date_limite_remise_offres', 'statut']
    for col_name in important_cols:
        if col_name in df_ordered.columns:
            col_idx = df_ordered.columns.get_loc(col_name) + 1
            col_letter = get_column_letter(col_idx)
            # Ajouter un fond légèrement différent pour les colonnes importantes
            for row in range(1, len(df) + 2):
                cell = ws[f'{col_letter}{row}']
                if row > 1:  # Pas l'en-tête
                    current_fill = cell.fill.start_color.rgb if cell.fill.start_color else '00000000'
                    if current_fill == '00F8F9FA' or current_fill == '00000000':
                        cell.fill = PatternFill(start_color='EDF2F7', end_color='EDF2F7', fill_type='solid')
    
    # Sauvegarde
    wb.save(output_path)
    print(f"\n✅ Fichier Excel généré : {output_path}")
    
    # Statistiques
    print(f"\n📈 Résumé :")
    print(f"   - Total marchés : {len(df)}")
    if 'statut' in df.columns:
        for status, count in df['statut'].value_counts().items():
            print(f"   - {status} : {count}")
    
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description='Convertit un CSV de marchés publics en Excel formaté',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples d'utilisation :
  python csv_to_excel_formatted.py input.csv
  python csv_to_excel_formatted.py data.csv -o output.xlsx
  python csv_to_excel_formatted.py final-v3-consolidated-classified-rule.csv
        """
    )
    
    parser.add_argument('input', help='Chemin vers le fichier CSV source')
    parser.add_argument('-o', '--output', help='Chemin du fichier Excel de sortie (optionnel)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.input):
        print(f"❌ Erreur : Le fichier {args.input} n'existe pas.")
        return 1
    
    try:
        output = convert_csv_to_excel(args.input, args.output)
        print(f"\n🎉 Conversion terminée avec succès !")
        return 0
    except Exception as e:
        print(f"❌ Erreur lors de la conversion : {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    exit(main())
