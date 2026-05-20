#!/usr/bin/env python3
"""
Script de complétion exhaustive du tableau comparatif RC.
Extraction déterministe sans LLM - conformément à la politique projet.
"""

import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import os
from datetime import datetime
import json

# Configuration
PDF_DIR = "/home/michka/Documents/0-AO-DCE/public/rc"
OUTPUT_DIR = "/home/michka/Documents/0-AO-DCE/data/output"
CSV_INPUT = f"{OUTPUT_DIR}/tableau_comparatif_RC.csv"
EXCEL_OUTPUT = f"{OUTPUT_DIR}/tableau_comparatif_RC_completed.xlsx"

# Mapping des fichiers PDF aux colonnes du tableau
PDF_MAPPING = {
    "RC chatbot.pdf": "2026-CHATBOT",
    "AE20260004_PORTAILS_HISI V2_RC_V1.0.pdf": "DAF_2026_000243",  # Banque de France - Portails HISI V2
    "RC PHASE CANDIDATURES.pdf": "MS26084",  # SYANE - Missions d'infogérance
    "2600006 - SPL -  IT Réseau Cloud Sécurité - RC .pdf": "2600006",
    "26A0133001_ RC.pdf": "26A0133001",
    "B26-01107-LS_RC.pdf": "B26-01107-LS",
    "26910A RC.pdf": "26910A",
    "1. 2026A0239_RC.pdf": "2026A0239",
}

# Fichiers MD associés (extraction déjà faite)
MD_FILES = {
    "AE20260004_PORTAILS_HISI-V2_RC_V1.0.md": None,
    "B26-01107-LS_RC.md": None,
    "RC-PHASE-CANDIDATURES.md": None,
}


def extract_text_from_pdf(pdf_path):
    """Extrait le texte complet d'un PDF page par page."""
    text_by_page = []
    full_text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                page_text = page.extract_text() or ""
                text_by_page.append({
                    'page': i,
                    'text': page_text
                })
                full_text += f"\n--- PAGE {i} ---\n{page_text}"
    except Exception as e:
        print(f"Erreur extraction {pdf_path}: {e}")
        return [], ""
    return text_by_page, full_text


def extract_from_markdown(md_path):
    """Extrait les informations d'un fichier Markdown."""
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        print(f"Erreur lecture MD {md_path}: {e}")
        return ""


def find_pattern(text, patterns, context_chars=200):
    """Recherche déterministe de motifs dans le texte."""
    results = []
    text_lower = text.lower()
    for pattern_name, pattern_list in patterns.items():
        for pattern in pattern_list:
            matches = list(re.finditer(pattern, text, re.IGNORECASE | re.DOTALL))
            for match in matches:
                start = max(0, match.start() - context_chars)
                end = min(len(text), match.end() + context_chars)
                context = text[start:end].replace('\n', ' ').strip()
                results.append({
                    'pattern_name': pattern_name,
                    'match': match.group(0),
                    'context': context[:300],
                    'position': match.start()
                })
    return results


def parse_date_limit(text):
    """Extraction déterministe des dates limites."""
    patterns = [
        r'Date\s+(?:et\s+heure\s+)?limite.*?reception.*?des.*?offres?[\s:]+(\d{1,2}[/-]\d{1,2}[/-]\d{4}).*?(\d{1,2}[:h]\d{2})?',
        r'DLRO.*?[:\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{4}).*?(\d{1,2}[:h]\d{2})?',
        r'limite.*?reception.*?candidatures?[\s:]+(\d{1,2}[/-]\d{1,2}[/-]\d{4}).*?(\d{1,2}[:h]\d{2})?',
        r'date\s+limite.*?d[eé]p[oô]t[\s:]+(\d{1,2}[/-]\d{1,2}[/-]\d{4}).*?(\d{1,2}[:h]\d{2})?',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            date = match.group(1)
            heure = match.group(2) if match.group(2) else "12:00"
            return f"{date} {heure}", match.group(0)[:300]
    return "", ""


def parse_procedure(text):
    """Extraction déterministe de la procédure."""
    patterns = [
        r'(appel\s+d\'offres\s+ouverte?)',
        r'(appel\s+d\'offres\s+restreinte?)',
        r'(proc[eé]dure\s+avec\s+n[eé]gociation)',
        r'(proc[eé]dure\s+adapt[eé]e?)',
        r'(accord-cadre)',
        r'(march[eé]\s+([àa]\s+tranches?|par\s+tranches?))',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1), match.group(0)
    return "", ""


def parse_duree(text):
    """Extraction déterministe de la durée."""
    patterns = [
        r'dur[eé]e\s+(?:du\s+march[eé])?[\s:]+(\d+)\s*(mois|an|ans?)(?:.*?reconduction\s+(\d+)\s*(?:an|ans?|mois)?)?',
        r'(\d+)\s*(mois|an|ans?)\s+ferme.*?\+\s*(\d+)\s*reconduction',
        r'dur[eé]e.*?[:\s]+(\d+)\s*(mois|an|ans?)',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            return match.group(0)[:200], match.group(0)
    return "", ""


def parse_montant(text):
    """Extraction déterministe des montants."""
    patterns = [
        r'montant\s+(?:maximum|max|plafond)[\s:]+([\d\s.,]+)\s*[€\$£]?(?:\s*HT)?',
        r'plafond[\s:]+([\d\s.,]+)\s*[€\$£]?(?:\s*HT)?',
        r'estimation[\s:]+([\d\s.,]+)\s*[€\$£]?(?:\s*HT)?',
        r'montant\s+total[\s:]+([\d\s.,]+)\s*[€\$£]?(?:\s*HT)?',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            montant = match.group(1).replace(' ', '').replace(',', '.')
            return f"{match.group(0)[:100]} HT", match.group(0)
    return "", ""


def parse_critere_attribution(text):
    """Extraction déterministe des critères d'attribution."""
    patterns = [
        r'crit[eè]res?\s+d\'attribution[\s:]+(.{0,500}?)(?=\n\d|\n[A-Z]|$)',
        r'pond[eé]ration[\s:]+(.{0,500}?)(?=\n\d|\n[A-Z]|$)',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            return match.group(1).strip()[:300], match.group(0)
    return "", ""


def parse_allotissement(text):
    """Extraction déterministe de l'allotissement."""
    patterns_oui = [
        r'alloti[\s:]+oui',
        r'(\d+)\s+lots?',
        r'recoupage\s+en\s+(\d+)\s+lots?',
    ]
    patterns_non = [
        r'alloti[\s:]+non',
        r'non\s+alloti',
        r'en\s+un\s+seul\s+lot',
    ]
    for pattern in patterns_oui:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return "Oui", match.group(0)
    for pattern in patterns_non:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return "Non", match.group(0)
    return "", ""


def parse_documents_candidature(text):
    """Extraction des documents de candidature demandés."""
    docs = []
    patterns = {
        'DC1': r'DC1|lettre\s+de\s+candidature',
        'DC2': r'DC2|d[eé]claration\s+du\s+candidat',
        'DUME': r'DUME',
        'DC4': r'DC4',
        'attestations': r'attestation.*?sociale|attestation.*?fiscale|attestation.*?r[gé]glementaire',
        'certifications': r'certification|qualification|ISO\s*\d+|LABEL|QUALIBAT',
        'r[eé]f[eé]rences': r'r[eé]f[eé]rence.*?prestation|r[eé]f[eé]rence.*?client',
        'CA': r'chiffre\s+d\'affaires|C\.A\.?|CA\s+des\s+soci[eé]t[eé]s',
    }
    for doc_type, pattern in patterns.items():
        if re.search(pattern, text, re.IGNORECASE):
            docs.append(doc_type)
    if docs:
        return "; ".join(docs), "Documents trouvés: " + ", ".join(docs)
    return "", ""


def parse_documents_offre(text):
    """Extraction des documents d'offre demandés."""
    docs = []
    patterns = {
        'CCTP': r'CCTP|cahier\s+des\s+clauses\s+techniques',
        'CCAP': r'CCAP|cahier\s+des\s+clauses\s+administratives',
        'CCAG': r'CCAG|cahier\s+des\s+clauses\s+g[eé]n[eé]rales',
        'BPU': r'BPU|bordereau\s+des\s+prix',
        'DPGF': r'DPGF|d[eé]composition\s+du\s+prix',
        'AE': r'acte\s+d\'engagement|AE',
        'm[eé]moire_technique': r'm[eé]moire\s+technique|offre\s+technique',
        'offre_financi[eè]re': r'offre\s+financi[eè]re|offre\s+commerciale',
        'DQE': r'DQE|d[eé]tail\s+quantitatif\s+estimatif',
    }
    for doc_type, pattern in patterns.items():
        if re.search(pattern, text, re.IGNORECASE):
            docs.append(doc_type)
    if docs:
        return "; ".join(docs), "Documents trouvés: " + ", ".join(docs)
    return "", ""


def parse_souverainete(text):
    """Extraction des exigences de souveraineté/RGPD."""
    patterns = [
        r'souverainet[eé]|h[eé]bergement\s+fran[cç]ais|donn[eé]es\s+en\s+france',
        r'RGPD|RSSI|s[eé]curit[eé]\s+des\s+syst[eè]mes?\s+d\'information',
        r'chiffrement|chiffrage|cryptage',
        r'localisation\s+des\s+donn[eé]es|h[eé]bergement\s+(?:europ[eé]en|UE)',
    ]
    found = []
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            found.extend(matches[:2])
    if found:
        unique_found = list(set([f.lower() for f in found]))
        return "; ".join(unique_found)[:300], "Exigences trouvées"
    return "", ""


def parse_soustraitance(text):
    """Extraction des exigences de sous-traitance."""
    patterns = [
        r'sous-traitance\s+(autoris[eé]e?|interdite?|possible)',
        r'taux\s+de\s+sous-traitance|plafond\s+de\s+sous-traitance',
        r'agrément\s+des\s+sous-traitants',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            return match.group(0)[:200], match.group(0)
    return "", ""


def parse_visite(text):
    """Extraction des modalités de visite."""
    patterns_oui = [
        r'visite\s+(?:obligatoire|requise|requise)',
        r'visite\s+des\s+lieux\s+obligatoire',
    ]
    patterns_non = [
        r'visite\s+non\s+obligatoire',
        r'pas\s+de\s+visite',
    ]
    for pattern in patterns_oui:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return "Oui - Obligatoire", match.group(0)
    for pattern in patterns_non:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return "Non", match.group(0)
    match = re.search(r'visite\s+(?:des\s+lieux|sur\s+place)', text, re.IGNORECASE)
    if match:
        return "Facultative", match.group(0)
    return "", ""


def parse_variantes(text):
    """Extraction des modalités de variantes."""
    patterns_autorise = [
        r'variants?\s+autoris[eé]e?s?(?:\s*:?\s*oui)?',
        r'variantes?\s+possible',
    ]
    patterns_non = [
        r'variants?\s+non\s+autoris[eé]e?s?',
        r'pas\s+de\s+variante',
    ]
    for pattern in patterns_autorise:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return "Oui", match.group(0)
    for pattern in patterns_non:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return "Non", match.group(0)
    return "", ""


def parse_delai_validite(text):
    """Extraction du délai de validité des offres."""
    patterns = [
        r'd[eé]lai\s+de\s+validit[eé].*?(\d+)\s*(jours?|mois?)',
        r'validit[eé].*?offres?.*?(\d+)\s*(jours?|mois?)',
        r'offre\s+valable.*?(\d+)\s*(jours?|mois?)',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            return f"{match.group(1)} {match.group(2)}", match.group(0)
    return "", ""


def parse_plateforme(text):
    """Extraction de la plateforme de dépôt."""
    patterns = [
        r'plateforme[\s:]+([^\n]{3,100})',
        r'd[eé]p[oô]t\s+sur\s+([^\n]{3,100})',
        r'(PLACE|e-march[eé]spublics|marches-publics\.info|safetender|mégalis|aws-mp\.fr|d[eé]mat\.acoss\.fr)',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip()[:100], match.group(0)
    return "", ""


def parse_signature_electronique(text):
    """Extraction des modalités de signature électronique."""
    patterns = [
        r'signature\s+[eé]lectronique',
        r'certificat\s+[eé]lectronique',
        r'format\s+accept[eé].*?(PDF|XML|ASiC)',
    ]
    found = []
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            found.append(match.group(0))
    if found:
        return "; ".join(found)[:200], "Signature électronique trouvée"
    return "", ""


def parse_sla(text):
    """Extraction des SLA/GTR/GTI."""
    patterns = [
        r'SLA|garantie\s+de\s+temps|GTR|GTI|GTS',
        r'temps\s+d\'intervention|d[eé]lai\s+d\'intervention',
        r'disponibilit[eé].*?(\d+)%',
        r'niveau\s+de\s+service|service\s+level',
    ]
    found = []
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            found.append(match.group(0)[:100])
    if found:
        return "; ".join(found[:3]), "SLA/GTR trouvé"
    return "", ""


def parse_certifications(text):
    """Extraction des certifications requises."""
    patterns = [
        r'ISO\s*27001|ISO\s*9001|ISO\s*14001|ISO\s*22301',
        r'QUALIBAT|OPQIBI|CERTIFER|LABEL|certification',
        r'SecNumCloud|ANSSI|CSPN',
    ]
    found = []
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            found.extend(matches)
    if found:
        return "; ".join(list(set(found)))[:200], "Certifications trouvées"
    return "", ""


def parse_reversibilite(text):
    """Extraction des obligations de reversibilité."""
    patterns = [
        r'reversibilit[eé]|plan\s+de\s+reversibilit[eé]',
        r'reprise\s+des\s+donn[eé]es|transfert\s+des\s+donn[eé]es',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(0)[:200], match.group(0)
    return "", ""


def parse_insertion_sociale(text):
    """Extraction des clauses d'insertion sociale."""
    patterns = [
        r'clause\s+d\'insertion|insertion\s+sociale|IAH|handicap|ESAT',
        r'emploi\s+de\s+travailleurs?\s+handicap[eé]s?',
        r'quata\s+handicap|obligation\s+d\'emploi',
    ]
    found = []
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            found.append(match.group(0)[:100])
    if found:
        return "; ".join(found), "Insertion sociale trouvée"
    return "", ""


def extract_all_data_from_pdf(pdf_filename):
    """Extraction complète des données d'un PDF."""
    pdf_path = os.path.join(PDF_DIR, pdf_filename)
    text_by_page, full_text = extract_text_from_pdf(pdf_path)
    
    if not full_text:
        return None, []
    
    # Extraire toutes les informations
    extraction = {
        'date_limite': parse_date_limit(full_text),
        'procedure_detail': parse_procedure(full_text),
        'duree': parse_duree(full_text),
        'montant_max': parse_montant(full_text),
        'critere_attribution': parse_critere_attribution(full_text),
        'allotissement': parse_allotissement(full_text),
        'documents_candidature': parse_documents_candidature(full_text),
        'documents_offre': parse_documents_offre(full_text),
        'souverainete_rgpd': parse_souverainete(full_text),
        'soustraitance': parse_soustraitance(full_text),
        'visite': parse_visite(full_text),
        'variantes': parse_variantes(full_text),
        'delai_validite': parse_delai_validite(full_text),
        'plateforme': parse_plateforme(full_text),
        'signature_electronique': parse_signature_electronique(full_text),
        'sla_gtr': parse_sla(full_text),
        'certifications': parse_certifications(full_text),
        'reversibilite': parse_reversibilite(full_text),
        'insertion_sociale': parse_insertion_sociale(full_text),
    }
    
    # Créer les logs avec sources
    logs = []
    for field, (value, source) in extraction.items():
        if value:
            logs.append({
                'fichier': pdf_filename,
                'champ': field,
                'valeur': value[:200],
                'source_extrait': source[:300] if source else ""
            })
    
    # Simplifier l'extraction pour le retour (juste les valeurs)
    simple_extraction = {k: (v[0] if v[0] else "") for k, v in extraction.items()}
    
    return simple_extraction, logs


def load_existing_csv():
    """Charge le CSV existant et retourne un DataFrame."""
    try:
        df = pd.read_csv(CSV_INPUT, encoding='utf-8')
        return df
    except Exception as e:
        print(f"Erreur chargement CSV: {e}")
        return None


def transpose_and_prepare_dataframe(df):
    """Transpose le DataFrame pour avoir les RC en lignes."""
    # Le CSV est actuellement en format : Critères en lignes, RC en colonnes
    # On veut : RC en lignes, Critères en colonnes
    
    df_transposed = df.set_index('Critère').T
    df_transposed.index.name = 'Reference_RC'
    df_transposed = df_transposed.reset_index()
    
    return df_transposed


def add_new_columns(df):
    """Ajoute les nouvelles colonnes au DataFrame."""
    new_columns = [
        'Procedure_detaillee',
        'Criteres_detailles',
        'Documents_candidature_obligatoires',
        'Documents_offre_obligatoires',
        'Exigences_souverainete_RGPD',
        'Soustraitance_conditions',
        'Visite_site_obligatoire',
        'Variantes_autorisees',
        'Delai_validite_offres',
        'Plateforme_depot',
        'Signature_electronique',
        'SLA_GTR_GTI',
        'Certifications_requises',
        'Reversibilite',
        'Insertion_sociale',
        'References_source'
    ]
    
    for col in new_columns:
        if col not in df.columns:
            df[col] = ""
    
    return df


def process_all_pdfs():
    """Traite tous les PDF et retourne les extractions et logs."""
    all_extractions = {}
    all_logs = []
    
    for pdf_file in PDF_MAPPING.keys():
        print(f"Extraction de {pdf_file}...")
        extraction, logs = extract_all_data_from_pdf(pdf_file)
        if extraction:
            col_name = PDF_MAPPING[pdf_file]
            if col_name:
                all_extractions[col_name] = extraction
                # Ajouter le fichier source aux références
                extraction['references_source'] = pdf_file
        if logs:
            all_logs.extend(logs)
    
    return all_extractions, all_logs


def fill_dataframe_with_extractions(df, extractions):
    """Remplit le DataFrame avec les données extraites."""
    column_mapping = {
        'Procedure_detaillee': 'procedure_detail',
        'Criteres_detailles': 'critere_attribution',
        'Documents_candidature_obligatoires': 'documents_candidature',
        'Documents_offre_obligatoires': 'documents_offre',
        'Exigences_souverainete_RGPD': 'souverainete_rgpd',
        'Soustraitance_conditions': 'soustraitance',
        'Visite_site_obligatoire': 'visite',
        'Variantes_autorisees': 'variantes',
        'Delai_validite_offres': 'delai_validite',
        'Plateforme_depot': 'plateforme',
        'Signature_electronique': 'signature_electronique',
        'SLA_GTR_GTI': 'sla_gtr',
        'Certifications_requises': 'certifications',
        'Reversibilite': 'reversibilite',
        'Insertion_sociale': 'insertion_sociale',
        'References_source': 'references_source',
    }
    
    for idx, row in df.iterrows():
        ref_rc = row['Reference_RC']
        if ref_rc in extractions:
            ext = extractions[ref_rc]
            for df_col, ext_key in column_mapping.items():
                if ext_key in ext and ext[ext_key]:
                    df.at[idx, df_col] = ext[ext_key]
    
    return df


def create_logs_sheet(logs):
    """Crée le DataFrame pour la feuille Logs & Sources."""
    if not logs:
        # Créer au moins une entrée avec la méthode
        logs = [{
            'fichier': 'METHODOLOGIE',
            'champ': 'Rappel',
            'valeur': 'Extraction déterministe sans LLM - conformément à la politique projet LLM OFF',
            'source_extrait': 'Règle : exclusivement utiliser les données des RC joints. Valeurs non présentes = cellules vides (option B).'
        }]
    
    df_logs = pd.DataFrame(logs)
    return df_logs


def create_missing_summary(df):
    """Crée la feuille Résumé des manques."""
    new_columns = [
        'Procedure_detaillee',
        'Criteres_detailles',
        'Documents_candidature_obligatoires',
        'Documents_offre_obligatoires',
        'Exigences_souverainete_RGPD',
        'Soustraitance_conditions',
        'Visite_site_obligatoire',
        'Variantes_autorisees',
        'Delai_validite_offres',
        'Plateforme_depot',
        'Signature_electronique',
        'SLA_GTR_GTI',
        'Certifications_requises',
        'Reversibilite',
        'Insertion_sociale',
    ]
    
    missing_data = []
    for idx, row in df.iterrows():
        ref_rc = row['Reference_RC']
        missing_fields = []
        for col in new_columns:
            if not row.get(col) or row.get(col) == "":
                missing_fields.append(col)
        
        missing_data.append({
            'Reference_RC': ref_rc,
            'Nombre_champs_vides': len(missing_fields),
            'Champs_non_precises': "; ".join(missing_fields) if missing_fields else "Aucun - RC complet"
        })
    
    return pd.DataFrame(missing_data)


def save_to_excel(df_main, df_logs, df_missing):
    """Sauvegarde les DataFrames dans un fichier Excel formaté."""
    with pd.ExcelWriter(EXCEL_OUTPUT, engine='openpyxl') as writer:
        # Feuille principale
        df_main.to_excel(writer, sheet_name='Comparatif RC', index=False)
        
        # Feuille Logs & Sources
        df_logs.to_excel(writer, sheet_name='Logs & Sources', index=False)
        
        # Feuille Résumé manques
        df_missing.to_excel(writer, sheet_name='Resume manques', index=False)
    
    # Formatage avec openpyxl
    wb = openpyxl.load_workbook(EXCEL_OUTPUT)
    
    # Formater la feuille principale
    ws_main = wb['Comparatif RC']
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for cell in ws_main[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Ajuster la largeur des colonnes
    for column in ws_main.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_main.column_dimensions[column_letter].width = adjusted_width
    
    # Formater les autres feuilles
    for sheet_name in ['Logs & Sources', 'Resume manques']:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
    
    wb.save(EXCEL_OUTPUT)
    print(f"Fichier Excel sauvegardé: {EXCEL_OUTPUT}")


def main():
    """Fonction principale."""
    print("=" * 60)
    print("COMPLETION TABLEAU COMPARATIF RC")
    print("Mode: Extraction déterministe sans LLM (Politique LLM OFF)")
    print("=" * 60)
    
    # 1. Charger le CSV existant
    print("\n[1/5] Chargement du CSV existant...")
    df = load_existing_csv()
    if df is None:
        print("Erreur: Impossible de charger le CSV")
        return
    print(f"CSV chargé: {len(df)} lignes, {len(df.columns)} colonnes")
    
    # 2. Transposer pour avoir les RC en lignes
    print("\n[2/5] Transposition des données...")
    df_transposed = transpose_and_prepare_dataframe(df)
    print(f"Données transposées: {len(df_transposed)} RC")
    
    # 3. Ajouter les nouvelles colonnes
    print("\n[3/5] Ajout des nouvelles colonnes...")
    df_transposed = add_new_columns(df_transposed)
    print(f"Colonnes ajoutées: {[c for c in df_transposed.columns if 'Reference' not in c and len(df_transposed.columns) > 15]}")
    
    # 4. Extraire les données des PDF
    print("\n[4/5] Extraction des données des PDF...")
    extractions, logs = process_all_pdfs()
    print(f"PDF traités: {len(extractions)}")
    print(f"Logs générés: {len(logs)}")
    
    # 5. Remplir le DataFrame
    print("\n[5/5] Remplissage du tableau...")
    df_filled = fill_dataframe_with_extractions(df_transposed, extractions)
    
    # 6. Créer les feuilles annexes
    df_logs = create_logs_sheet(logs)
    df_missing = create_missing_summary(df_filled)
    
    # 7. Sauvegarder
    save_to_excel(df_filled, df_logs, df_missing)
    
    print("\n" + "=" * 60)
    print("TRAITEMENT TERMINE")
    print(f"Fichier généré: {EXCEL_OUTPUT}")
    print("=" * 60)
    
    # Résumé
    print("\nRésumé des extractions par RC:")
    for ref, ext in extractions.items():
        filled_count = sum(1 for v in ext.values() if v)
        total_count = len(ext)
        print(f"  - {ref}: {filled_count}/{total_count} champs remplis")


if __name__ == "__main__":
    main()
